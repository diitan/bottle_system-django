from django.conf import settings
from django.utils.dateparse import parse_date
from django.db.models import Count
from django.shortcuts import render, redirect
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from django.http import HttpResponse
from datetime import datetime
from django.contrib.auth.decorators import login_required, user_passes_test
from orders.models import Order
from django.contrib.auth import logout
from django.core.paginator import Paginator
from django.http import JsonResponse
from django.views.decorators.http import require_POST

def home(request):
    return render(request, 'core/home.html')

# --- admin logout ---
def admin_logout(request):
    logout(request)
    return redirect("login")

# --- admin login & dashboard ---
def is_admin(user):
    return user.is_staff or user.is_superuser
@login_required
@user_passes_test(is_admin)
def admin_dashboard(request):
    # filter params (GET)
    #start_raw = request.GET.get("start_date", "").strip()
    start_raw = (request.GET.get("start_date") or "").strip()
    #end_raw = request.GET.get("end_date", "").strip()
    end_raw = (request.GET.get("end_date") or "").strip()

    qs = Order.objects.all().order_by("-created_at")

    start_date = None
    end_date = None
    if start_raw:
        try:
            #start_date = datetime.strptime(start_raw, "%Y-%m-%d").date()
            start_date = datetime.strptime(start_raw, "%d/%m/%Y")
        except Exception:
            start_date = None
    if end_raw:
        try:
            #end_date = datetime.strptime(end_raw, "%Y-%m-%d").date()
            end_date = datetime.strptime(end_raw, "%d/%m/%Y")
        except Exception:
            end_date = None

    if start_date:
        qs = qs.filter(created_at__date__gte=start_date)
    if end_date:
        qs = qs.filter(created_at__date__lte=end_date)

    # statistics
    total = qs.count()
    processed = qs.filter(is_paid=True).count()
    pending = qs.filter(is_paid=False).count()

    # Build simple list to show (no pagination here — can add later)
    orders = qs.values("id", "full_name", "created_at", "total_price", "is_paid")
    orders_list = Order.objects.all().order_by('-created_at')
    paginator = Paginator(orders_list, 10)
    page_number = request.GET.get('page')
    orders = paginator.get_page(page_number)

    context = {
        "orders": orders,
        "total": total,
        "processed": processed,
        "pending": pending,
        "start_date": start_raw,
        "end_date": end_raw,
    }
    return render(request, "admin/dashboard.html", context)

# --- export csv (Excel can open) ---
@login_required
@user_passes_test(is_admin)
def admin_export_excel(request):
    orders = Order.objects.all().order_by('-created_at')

    wb = Workbook()
    ws = wb.active
    ws.title = "Danh sách đơn hàng"

    # ===== TITLE =====
    ws.merge_cells("A1:F1")
    ws["A1"] = "BÁO CÁO DANH SÁCH ĐƠN HÀNG"
    ws["A1"].font = Font(size=16, bold=True)
    ws["A1"].alignment = Alignment(horizontal="center")

    ws.merge_cells("A2:F2")
    ws["A2"] = f"Ngày xuất: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
    ws["A2"].alignment = Alignment(horizontal="center")

    # ===== HEADER =====
    headers = ["STT", "ID", "Họ tên", "Ngày đặt", "Tổng tiền", "Trạng thái"]
    ws.append(headers)

    header_fill = PatternFill(start_color="4F81BD",
                              end_color="4F81BD",
                              fill_type="solid")

    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=3, column=col)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal="center")
        cell.fill = header_fill

    # ===== DATA =====
    row_num = 4
    for index, order in enumerate(orders, start=1):
        ws.cell(row=row_num, column=1, value=index)
        ws.cell(row=row_num, column=2, value=order.id)
        ws.cell(row=row_num, column=3, value=order.full_name)
        ws.cell(row=row_num, column=4, value=order.created_at.strftime("%d-%m-%Y %H:%M"))
        ws.cell(row=row_num, column=5, value=order.total_price)
        ws.cell(row=row_num, column=6, value="Đã thanh toán" if order.is_paid else "Chưa thanh toán")

        ws.cell(row=row_num, column=5).number_format = '#,##0 "VNĐ"'
        ws.cell(row=row_num, column=5).alignment = Alignment(horizontal="right")

        row_num += 1

    # ===== AUTO WIDTH =====
    for col in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        ws.column_dimensions[col_letter].width = max_length + 4

    # ===== RESPONSE =====
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="bao_cao_don_hang.xlsx"'

    wb.save(response)
    return response

@require_POST
def delete_order(request, order_id):
    try:
        order = Order.objects.get(id=order_id)
        order.delete()
        return JsonResponse({"success": True})
    except Order.DoesNotExist:
        return JsonResponse({"success": False})
